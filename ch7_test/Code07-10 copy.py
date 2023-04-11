import openpyxl
from copy import copy

workbook = openpyxl.load_workbook('Excel/singer.xlsx')
#엑셀 파일 시트 이름 리스트
wsheetList = workbook.sheetnames
#기본으로 빈 엑셀파일(워크북)
outWorkbook = openpyxl.Workbook()
# 기본으로 생성된 시트를 일단 제거
outWorkbook.remove(outWorkbook['Sheet']) 

# 워크시트 이름을 하나씩 꺼내서 작업.
for wsName in wsheetList :
    #workbook -> 엑셀 파일 해당하는 워크시트 하나 선택.
    worksheet = workbook[wsName]
    # 새롭게 만들 워크시트의 이름 -> 이름 만들기 작업.
    outSheet = outWorkbook.create_sheet('New_' + wsName)
    #openpyxl 여기서는 0행이 아니라, 1행부터 작업을 함.
    # 그래서 끝행도 하나가 더 추가 됩니다. 즉, 한칸씩 값이 증가.
    totRow = 0 # 출력용 엑셀의 행번호
    for row in range(1, worksheet.max_row+1) :
        if row != 1 :
            if int(worksheet.cell(row=row, column=5).value) <= 165 :
                continue
        totRow += 1
        outSheet.row_dimensions[row].height = worksheet.row_dimensions[row].height
        for col in range(1, worksheet.max_column+1) :
            outSheet.column_dimensions[chr(ord('A')+col-1)].width \
                = worksheet.column_dimensions[chr(ord('A')+col-1)].width
            inCell = worksheet.cell(row=row, column=col)
            outCell = outSheet.cell(row=totRow, column=col, value= inCell.value)
            if inCell.has_style:
                outCell.font = copy(inCell.font)
                outCell.border = copy(inCell.border)
                outCell.fill = copy(inCell.fill)
                outCell.number_format = copy(inCell.number_format)
                outCell.alignment = copy(inCell.alignment)
outWorkbook.save('Excel/singer_copy3.xlsx')
print("Save. OK~")

#     for row in range(1, worksheet.max_row+1) :
#         # 읽은 엑셀파일의 높이와, outsheet의 높이 같게.
#         outSheet.row_dimensions[row].height = worksheet.row_dimensions[row].height
#         #1행의 1열부터 작업시작
#         for col in range(1, worksheet.max_column+1) :
#             # ord('A') -> 숫자, chr(65) -> 문자.
#             outSheet.column_dimensions[chr(ord('A')+col-1)].width \
#                 = worksheet.column_dimensions[chr(ord('A')+col-1)].width
#             #현재 읽으려는 파일의 각 셀
#             inCell = worksheet.cell(row=row, column=col)
#             #출력하기 위한 각 셀
#             outCell = outSheet.cell(row=row, column=col, value= inCell.value)
#             if inCell.has_style:
#                 outCell.font = copy(inCell.font)
#                 outCell.border = copy(inCell.border)
#                 outCell.fill = copy(inCell.fill)
#                 outCell.number_format = copy(inCell.number_format)
#                 outCell.alignment = copy(inCell.alignment)


